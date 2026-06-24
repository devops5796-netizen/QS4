import pandas as pd
import re
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

INVALID_CHARS = re.compile(r'[\\/*?:\[\]]')

def clean_name(name: str, max_len: int = 31) -> str:
    return INVALID_CHARS.sub('_', str(name))[:max_len]

def style_header(ws):
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def write_manufacturer_excel(manufacturer: str, df_mfr: pd.DataFrame, output_dir: Path):
    filepath = output_dir / f"{clean_name(manufacturer)}.xlsx"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:

        # Sheet لكل موديل (الصنف)
        for model in sorted(df_mfr["الصنف"].unique(), key=str):
            df_model = df_mfr[df_mfr["الصنف"] == model].reset_index(drop=True)
            sheet_name = clean_name(str(model))
            df_model.to_excel(writer, sheet_name=sheet_name, index=False)
            style_header(writer.sheets[sheet_name])

    return filepath

def run(input_csv: str = "cars_for_sale.csv", output_dir: str = "excel_by_manufacturer"):
    print(f"Reading {input_csv}...")
    df = pd.read_excel(input_csv)
    print(f"Total rows: {len(df)}")

    out = Path(output_dir)
    out.mkdir(exist_ok=True)

    df["المصنع"] = df["المصنع"].fillna("غير محدد")
    df["الصنف"]  = df["الصنف"].fillna("غير محدد")

    manufacturers = df["المصنع"].unique()
    print(f"Found {len(manufacturers)} manufacturers\n")

    results = []
    for mfr in sorted(manufacturers, key=str):
        df_mfr = df[df["المصنع"] == mfr].copy()
        filepath = write_manufacturer_excel(str(mfr), df_mfr, out)
        n_models = df_mfr["الصنف"].nunique()
        print(f"  ✓ {mfr}: {len(df_mfr)} rows | {n_models} models → {filepath.name}")
        results.append({"manufacturer": mfr, "rows": len(df_mfr), "models": n_models, "file": filepath.name})

    # Index file
    summary_df = pd.DataFrame(results).sort_values("rows", ascending=False)
    summary_path = out / "00_index.xlsx"
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Index", index=False)
        style_header(writer.sheets["Index"])

    print(f"\nIndex → {summary_path}")
    print(f"Done. {len(manufacturers)} Excel files in '{output_dir}/'")

if __name__ == "__main__":
    import sys
    csv_path = sys.argv[1] if len(sys.argv) > 1 else "cars_for_sale.csv"
    out_path  = sys.argv[2] if len(sys.argv) > 2 else "excel_by_manufacturer"
    run(csv_path, out_path)